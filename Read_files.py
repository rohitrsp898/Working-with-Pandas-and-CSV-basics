import os
import pandas as pd
from openpyxl import load_workbook
import csv

#all files
all_csv=[x for x in os.listdir("C:\\Users\\roprajap\\Desktop\\Python_Automation_IDEA\\AIML\\Data\\") if x.endswith(".csv")]

# Record_Counts=open("Records_Counts.csv",'w',newline='')
# writer_object = csv.writer(Record_Counts)

# workbook_name = 'output.xlsx'
# wb = load_workbook(workbook_name)
# sheet = wb.active

def panda_method(all_csv):
    # open/create spreadsheet in writer
    writer = pd.ExcelWriter('output.xlsx')
    tb_row_count = []


    for i in range(0,len(all_csv)):
        df = pd.read_csv("C:\\Users\\roprajap\\Desktop\\Python_Automation_IDEA\\AIML\\Data\\" + all_csv[i])

        # Count all records including null
        tb_row_count.append((all_csv[i][:-4],(df[df.columns[0]].count()+(df[df.columns[0]].isna().sum()))))

        # Get records from 1st col
        dup_records = df[df.iloc[:, :1].duplicated(keep=False)]
        Null_records = df[df.iloc[:, :1].isnull().any(axis=1)]

        # Duplicate records
        if not dup_records.empty:
            #dup_records.to_csv("dup_records_in_" + all_csv[i], index=False)
            dup_records.to_excel(writer, sheet_name='Dup '+all_csv[i][:-4]+" data", index=False)

        # Null records
        if not Null_records.empty:
            #Null_records.to_csv("NULL_records_in_"+all_csv[i], index=False)
            Null_records.to_excel(writer, sheet_name='Null '+all_csv[i][:-4]+" data",index=False)

    # Rows Count
    row_count=pd.DataFrame(tb_row_count,columns=["Table",'Row Count'])
    row_count.to_excel(writer, sheet_name="Row Count", index=False)
    writer.save()



def table_Columns(all_csv):
    workbook_name = 'output.xlsx'
    wb = load_workbook(workbook_name)

    # Worksheet for Table Details
    wb.create_sheet("Table Details")
    tb_records = wb["Table Details"]
    tb_records.append(["Tables", "Coulmns", "Primary Keys"])

    # Worksheet for PK details
    wb.create_sheet("PK")
    PK = wb["PK"]
    PK.append(["Tables", "Primary Keys", "Nullable", "Duplicates"])


    table_col = {}
    for i in range(0, len(all_csv)):
        df = pd.read_csv("C:\\Users\\roprajap\\Desktop\\Python_Automation_IDEA\\AIML\\Data\\" + all_csv[i])

        #Table, PK, Nullable, Duplicates
        #----------------------------------
        dup_rec = df[df.iloc[:, :1].duplicated(keep=False)]
        Null_rec = df[df.iloc[:, :1].isnull().any(axis=1)]


        if dup_rec.empty and Null_rec.empty:
            PK.append([all_csv[i][:-4], df.columns[:][0], "No", "NA"])
            #print([all_csv[i][:-4], df.columns[:][0], "No", "NA"])
        elif dup_rec.empty and not Null_rec.empty:
            PK.append([all_csv[i][:-4], df.columns[:][0], "Yes", "NA"])
            #print([all_csv[i][:-4], df.columns[:][0], "Yes", 'NA'])
        elif not dup_rec.empty and not Null_rec.empty:
            PK.append([all_csv[i][:-4], df.columns[:][0], "Yes", str(dup_rec.iloc[:, :1].values.tolist())])
            #print([all_csv[i][:-4], df.columns[:][0], "Yes", dup_rec.iloc[:, :1].values.tolist()])
        else:
            PK.append([all_csv[i][:-4], df.columns[:][0], "No", str(dup_rec.iloc[:, :1].values.tolist())])
            #print([all_csv[i][:-4], df.columns[:][0], "No", dup_rec.iloc[:, :1].values.tolist()])

        #--------------------------------------

        # Created dict with empty list values
        table_col.setdefault(all_csv[i][:-4], [])
        for col in df.columns:
            table_col[all_csv[i][:-4]].append(col)


    # table, all coulmns, PK
    for key, values in table_col.items():

        for i in values:
            if i == values[0]:

                tb_records.append([key, i, "PK"])
            else:
                if i == values[-1]:
                    tb_records.append(["", ""])
                else:
                    tb_records.append(["", i])

    wb.save(filename=workbook_name)


panda_method(all_csv)
table_Columns(all_csv)



